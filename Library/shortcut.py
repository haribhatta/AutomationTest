import openpyxl
from openpyxl.styles import PatternFill


def getTotalrows(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)


def getTotalcolumns(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)


def readData(file, sheetName, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum, column=columnnum).value


def writeData(file, sheetName, rownum, columnnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum, column=columnnum).value = data
    workbook.save(file)


def copyoriginal(original, duplicate):
    workbook = openpyxl.load_workbook(original)
    workbook.save(duplicate)


def fillGreenColor(file, sheetName, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    greenFill = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')
    sheet.cell(rownum, columnnum).fill = greenFill
    workbook.save(file)


def fillYellowColor(file, sheetName, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    yellowFill = PatternFill(start_color='FFFF00',
                            end_color='FFFF00',
                            fill_type='solid')
    sheet.cell(rownum, columnnum).fill = yellowFill
    workbook.save(file)
