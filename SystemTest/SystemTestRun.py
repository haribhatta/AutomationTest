import openpyxl
import pathlib
#theFile = openpyxl.load_workbook('SystemTestCases.xlsx')
theFile = openpyxl.load_workbook('SystemTestExecutionResultV1.48.8.xlsx')

allSheetNames = theFile.sheetnames
#path = 'D:\svn\DanpheEMR\SystemTest'

print("All sheet names {} " .format(theFile.sheetnames))

def find_specific_cell():
    for row in range(2, 59):
        for column in "B":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            cellstatus = "{}{}".format("D", row)
            #tcStatus = currentSheet[cell_name].value
            if  currentSheet[cell_name].value != "Passed":
                print(" I am executing no run & failed test cases")
                #cell_namer = "{}{}".format('A', row)
                cell_namer = "{}{}".format('A', row)
                print("Test Case Executing", currentSheet[cell_namer].value)
                tctorun = currentSheet[cell_namer].value
                base_dir = pathlib.PureWindowsPath(r'D:\svn\DanpheEMR\SystemTest')
                filenametorun = base_dir / tctorun
                print(filenametorun)
                try:
                    exec(open(filenametorun).read())
                    currentSheet[cell_name].value = "Passed"
                except Exception:
                    currentSheet[cell_name].value = "Failed"
                    currentSheet[cellstatus].value = currentSheet[cellstatus].value + 1
                    pass
                theFile.save('SystemTestExecutionResultV1.48.8.xlsx')
                print("Updating status")
                print(currentSheet)
                print(cell_name)
                print(currentSheet[cell_name].value)
                #return cell_name

def get_all_values_by_cell_letter(letter):
    for row in range(1, currentSheet.max_row + 1):
        for column in letter:
            cell_name = "{}{}".format(column, row)
            #print(cell_name)
            print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))

for sheet in allSheetNames:
    print("Current sheet name is {}" .format(sheet))
    currentSheet = theFile[sheet]
    specificCellLetter = (find_specific_cell())